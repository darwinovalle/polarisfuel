from celery import shared_task
from openpyxl import load_workbook
from pathlib import Path
import csv
import os

from stations.services.import_pipeline import run_import
from stations.services.parsing import REQUIRED_HEADERS

def _validate_and_parse_rows(rows):
    rows = list(rows)
    if not rows:
        raise ValueError("Input file is empty")

    headers = [str(header).strip() for header in rows[0].keys()]
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise ValueError(f"Missing headers: {', '.join(missing)}")

    return [
        {headers[index]: values for index, values in enumerate(row.values())}
        for row in rows
    ]


def _read_excel_rows(file_path: str):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("Excel file is empty")

    headers = [str(header).strip() if header is not None else "" for header in header_row]

    parsed_rows = []
    for values in rows_iter:
        row = {headers[i]: values[i] for i in range(len(headers))}
        parsed_rows.append(row)

    return _validate_and_parse_rows(parsed_rows)


def _read_csv_rows(file_path: str):
    with open(file_path, newline="", encoding="utf-8-sig") as csv_file:
        return _validate_and_parse_rows(csv.DictReader(csv_file))


def _read_rows(file_path: str):
    suffix = Path(file_path).suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(file_path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_excel_rows(file_path)
    raise ValueError("Unsupported file format. Use CSV or Excel (.xlsx, .xlsm, .xltx, .xltm)")

@shared_task(name="stations.tasks.run_daily_import")
def run_daily_import():
    source_path = os.getenv("EXCEL_SOURCE_PATH", "").strip()
    if not source_path:
        return {"status": "skipped", "reason": "EXCEL_SOURCE_PATH not set"}
    path = Path(source_path)
    if not path.exists():
        return {"status": "skipped", "reason": f"file not found: {source_path}"}

    rows = _read_rows(str(path))
    result = run_import(rows, source_filename=path.name)
    return {"status": "ok", "result": result}
